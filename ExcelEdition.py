from skyfield.api import load, wgs84
from datetime import date
import openpyxl
import matplotlib.pyplot as plt

wb=openpyxl.Workbook()
sheet= wb.active

dataAzi=[]
dataAl=[]
sat=[]

satellites = load.tle_file('gp.php')
print('Loaded', len(satellites), 'satellites')
by_number = {sat.name: sat for sat in satellites}
count=0
bluffton = wgs84.latlon(12.933851, 77.691874,920)

ts = load.timescale()
for k in range(100):
    t=ts.now()
    j=1
    for i in by_number:
            
            satellite = by_number[i]
            
            difference = satellite - bluffton
            topocentric = difference.at(t)
            alt, az, distance = topocentric.altaz()
            count=count+1
            if alt.degrees > 20:
                cn=sheet.cell(row=j,column=1)
                cn.value =float(alt.degrees)
                cn=sheet.cell(row=j,column=2)
                cn.value =float(az.degrees)
                cn=sheet.cell(row=j,column=3)
                cn.value =str(i)
                j=j+1

    wb.save("sam.xlsx")    
    path="sam.xlsx"
    wb_obj=openpyxl.load_workbook(path)
    sheetl=wb_obj.active
    cellaz = sheetl.cell(row=int(k/10+1),column=1)
    cellal= sheetl.cell(row=int(k/10+1),column=2)
    cellsat= sheetl.cell(row=int(k/10+1),column=3)
    dataAzi.append(cellaz.value)
    dataAl.append(cellal.value)
    sat.append(cellsat.value)
    #print(cellaz.value)
sat=list(set(sat))
x=list(range(len(dataAzi)))

plt.plot(x,dataAzi,label='AZIMUTH')

plt.plot(x,dataAl,label='ALTIDUDE')
plt.xlabel('TIME')
plt.ylabel('ANGLE')
plt.title('SATELLITE TRACKING GRAPH')
plt.ticklabel_format(style='plain', axis='both')
plt.legend()
plt.show()
for a in sat:
    print(a)